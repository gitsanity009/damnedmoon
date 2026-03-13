#!/usr/bin/env python3
"""
fix_choice_buttons.py — Add clickable choice buttons to the Game sheet.

The Game sheet is missing the Shape buttons (btnChoice1-5) that the VBA engine
expects. This script injects them directly into the .xlsm ZIP archive by:
  1. Creating a DrawingML drawing with 5 rounded-rectangle shapes
  2. Each shape has a `macro` attribute pointing to ChoiceClicked_N
  3. Wiring the drawing into the Game sheet (sheet1) relationships

Run:  python3 fix_choice_buttons.py
"""

import os
import shutil
import zipfile
import xml.etree.ElementTree as ET

WORKBOOK = "DamnedMoonv2.xlsm"
BACKUP = "DamnedMoonv2_backup.xlsm"

# ── Button layout constants (matching VBA's SetupGame) ──
# Buttons go at rows 25-29 (0-indexed: 24-28), columns B:C merged
# Using EMU (English Metric Units): 1 inch = 914400 EMU, 1 pt = 12700 EMU
# Excel default column width ≈ 64px, row height 24pt

# We need to position using twoCellAnchor with from/to cell references.
# Buttons span columns B-C (col index 1-2), one row each, rows 25-29 (0-indexed 24-28)

BUTTON_ROWS = [24, 25, 26, 27, 28]  # 0-indexed row numbers for rows 25-29
BTN_COL_FROM = 1   # Column B (0-indexed)
BTN_COL_TO = 3     # Column D (0-indexed, exclusive end approximation)
BTN_INSET_X = 25400    # ~2pt left inset
BTN_INSET_Y = 12700    # ~1pt top inset
BTN_OUTSET_X = 50800   # ~4pt right outset from col end
BTN_OUTSET_Y = 25400   # ~2pt bottom outset

# Colors in hex (RRGGBB for DrawingML is actually RRGGBB)
C_GOLD = "C9A227"       # RGB(201, 162, 39)
C_PANEL = "221A12"       # RGB(34, 26, 18)
C_BORDER = "3A2E22"      # RGB(58, 46, 34)

# Namespaces
NS_DRAWING = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
NS_A = "http://schemas.openxmlformats.org/drawingml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_SPREADSHEET = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
NS_CONTENT = "http://schemas.openxmlformats.org/package/2006/content-types"
NS_DRAWING_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"


def make_shape_xml(idx, row_idx, display_text="", visible=True):
    """Build XML for one rounded-rectangle shape button."""
    btn_name = f"btnChoice{idx}"
    macro_name = f"ChoiceClicked_{idx}"
    # Escape XML special characters in display text
    display_text = display_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("'", "&apos;").replace('"', "&quot;")

    shape = f"""
    <xdr:twoCellAnchor xmlns:xdr="{NS_DRAWING}" xmlns:a="{NS_A}">
      <xdr:from>
        <xdr:col>{BTN_COL_FROM}</xdr:col>
        <xdr:colOff>{BTN_INSET_X}</xdr:colOff>
        <xdr:row>{row_idx}</xdr:row>
        <xdr:rowOff>{BTN_INSET_Y}</xdr:rowOff>
      </xdr:from>
      <xdr:to>
        <xdr:col>{BTN_COL_TO}</xdr:col>
        <xdr:colOff>0</xdr:colOff>
        <xdr:row>{row_idx + 1}</xdr:row>
        <xdr:rowOff>0</xdr:rowOff>
      </xdr:to>
      <xdr:sp macro="{macro_name}" textlink="">
        <xdr:nvSpPr>
          <xdr:cNvPr id="{idx + 1}" name="{btn_name}"/>
          <xdr:cNvSpPr/>
        </xdr:nvSpPr>
        <xdr:spPr>
          <a:xfrm>
            <a:off x="0" y="0"/>
            <a:ext cx="0" cy="0"/>
          </a:xfrm>
          <a:prstGeom prst="roundRect">
            <a:avLst>
              <a:gd name="adj" fmla="val 8000"/>
            </a:avLst>
          </a:prstGeom>
          <a:solidFill>
            <a:srgbClr val="{C_PANEL}"/>
          </a:solidFill>
          <a:ln w="9525">
            <a:solidFill>
              <a:srgbClr val="{C_BORDER}"/>
            </a:solidFill>
          </a:ln>
        </xdr:spPr>
        <xdr:txBody>
          <a:bodyPr vertOverflow="clip" horzOverflow="clip" wrap="square"
                     lIns="152400" rIns="101600" tIns="25400" bIns="25400"
                     anchor="ctr"/>
          <a:lstStyle/>
          <a:p>
            <a:pPr algn="l"/>
            <a:r>
              <a:rPr lang="en-US" sz="1100" b="0" dirty="0">
                <a:solidFill>
                  <a:srgbClr val="{C_GOLD}"/>
                </a:solidFill>
                <a:latin typeface="Georgia"/>
                <a:cs typeface="Georgia"/>
              </a:rPr>
              <a:t>{display_text}</a:t>
            </a:r>
          </a:p>
        </xdr:txBody>
      </xdr:sp>
      <xdr:clientData/>
    </xdr:twoCellAnchor>"""
    return shape


def get_current_choices():
    """Read the current scene's choice texts from the workbook."""
    import openpyxl
    wb = openpyxl.load_workbook(WORKBOOK, keep_vba=True, data_only=True)
    ws_game = wb['Game']
    ws_scenes = wb['tbl_Scenes']

    scene_id = ws_game['E40'].value
    if not scene_id:
        return [""] * 5

    choices = []
    for r in range(2, ws_scenes.max_row + 1):
        if ws_scenes.cell(r, 1).value == scene_id:
            for i, col in enumerate([7, 11, 15, 19, 23], 1):
                text = ws_scenes.cell(r, col).value
                if text:
                    choices.append(f"{i}.  {text}")
                else:
                    choices.append("")
            break
    else:
        choices = [""] * 5

    # Pad to 5
    while len(choices) < 5:
        choices.append("")

    return choices


def build_drawing_xml():
    """Build the complete drawing1.xml content."""
    choices = get_current_choices()

    shapes = ""
    for i, row in enumerate(BUTTON_ROWS, 1):
        text = choices[i - 1]
        visible = bool(text)
        shapes += make_shape_xml(i, row, display_text=text, visible=visible)

    xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="{NS_DRAWING}"
           xmlns:a="{NS_A}"
           xmlns:r="{NS_R}">
{shapes}
</xdr:wsDr>"""
    return xml


def build_drawing_rels():
    """Build the _rels file for the drawing (empty, no external refs)."""
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_REL}"/>"""


def patch_workbook(src_path, dst_path):
    """Patch the .xlsm file to add the drawing with choice buttons."""

    with zipfile.ZipFile(src_path, 'r') as zin, \
         zipfile.ZipFile(dst_path, 'w', zipfile.ZIP_DEFLATED) as zout:

        # Track what we need to modify
        sheet1_rels_path = "xl/worksheets/_rels/sheet1.xml.rels"
        sheet1_path = "xl/worksheets/sheet1.xml"
        content_types_path = "[Content_Types].xml"
        drawing_path = "xl/drawings/drawing1.xml"
        drawing_rels_path = "xl/drawings/_rels/drawing1.xml.rels"

        # Check if sheet1 rels exists
        existing_files = set(zin.namelist())
        has_sheet1_rels = sheet1_rels_path in existing_files

        for item in zin.namelist():
            data = zin.read(item)

            if item == content_types_path:
                # Add content type for drawing
                data = patch_content_types(data, drawing_path)

            elif item == sheet1_path:
                # Add drawing reference to sheet1
                data = patch_sheet_xml(data)

            elif item == sheet1_rels_path:
                # Add drawing relationship
                data = patch_sheet_rels(data)
                has_sheet1_rels = True  # Mark as handled

            zout.writestr(item, data)

        # If sheet1 rels didn't exist, create it
        if not has_sheet1_rels:
            rels_xml = create_sheet_rels()
            zout.writestr(sheet1_rels_path, rels_xml)

        # Add the drawing file
        zout.writestr(drawing_path, build_drawing_xml())

        # Add drawing rels directory
        zout.writestr(drawing_rels_path, build_drawing_rels())


def patch_content_types(data, drawing_path):
    """Add drawing content type to [Content_Types].xml."""
    ET.register_namespace('', NS_CONTENT)
    root = ET.fromstring(data)

    # Check if drawing override already exists
    for override in root.findall(f'{{{NS_CONTENT}}}Override'):
        if override.get('PartName', '').endswith('drawing1.xml'):
            return data  # Already there

    # Add Override for the drawing
    override = ET.SubElement(root, f'{{{NS_CONTENT}}}Override')
    override.set('PartName', '/' + drawing_path)
    override.set('ContentType',
                 'application/vnd.openxmlformats-officedocument.drawing+xml')

    return ET.tostring(root, xml_declaration=True, encoding='UTF-8')


def patch_sheet_xml(data):
    """Add <drawing r:id="rIdDrawing1"/> to sheet1.xml."""
    # Register namespaces to preserve them
    namespaces = {
        '': NS_SPREADSHEET,
        'r': NS_R,
    }
    for prefix, uri in namespaces.items():
        ET.register_namespace(prefix, uri)

    # Parse and find all existing namespaces
    root = ET.fromstring(data)

    # Check if drawing element already exists
    for elem in root:
        if 'drawing' in elem.tag.lower():
            return data  # Already has a drawing

    # Find the right insertion point - drawing should go after certain elements
    # per the schema order. We'll append it at the end (before closing tag).
    drawing_elem = ET.SubElement(root, f'{{{NS_SPREADSHEET}}}drawing')
    drawing_elem.set(f'{{{NS_R}}}id', 'rIdDrawing1')

    return ET.tostring(root, xml_declaration=True, encoding='UTF-8')


def patch_sheet_rels(data):
    """Add drawing relationship to sheet1.xml.rels."""
    ET.register_namespace('', NS_REL)
    root = ET.fromstring(data)

    # Check if drawing rel already exists
    for rel in root:
        if 'drawing' in rel.get('Type', '').lower():
            return data

    rel = ET.SubElement(root, 'Relationship')
    rel.set('Id', 'rIdDrawing1')
    rel.set('Type', NS_DRAWING_REL)
    rel.set('Target', '../drawings/drawing1.xml')

    return ET.tostring(root, xml_declaration=True, encoding='UTF-8')


def create_sheet_rels():
    """Create a new sheet1.xml.rels with just the drawing relationship."""
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{NS_REL}">
  <Relationship Id="rIdDrawing1"
    Type="{NS_DRAWING_REL}"
    Target="../drawings/drawing1.xml"/>
</Relationships>"""


def update_cell_text():
    """Update the choice cell text (B25-B29) to match current scene choices."""
    import openpyxl as xl
    wb = xl.load_workbook(WORKBOOK, keep_vba=True)
    ws = wb['Game']
    ws_scenes = wb['tbl_Scenes']

    scene_id = ws['E40'].value
    choices = []
    if scene_id:
        for r in range(2, ws_scenes.max_row + 1):
            if ws_scenes.cell(r, 1).value == scene_id:
                for col in [7, 11, 15, 19, 23]:
                    text = ws_scenes.cell(r, col).value
                    choices.append(text or "")
                break

    # Pad to 5
    while len(choices) < 5:
        choices.append("")

    # Update cells
    for i, text in enumerate(choices):
        idx = i + 1
        cell_text = f"{idx}.  {text}" if text else None
        ws.cell(row=25 + i, column=2).value = cell_text

    wb.save(WORKBOOK)
    print(f"Updated cell text for {sum(1 for c in choices if c)} choices")


def main():
    if not os.path.exists(WORKBOOK):
        print(f"ERROR: {WORKBOOK} not found")
        return

    # Backup
    shutil.copy2(WORKBOOK, BACKUP)
    print(f"Backup saved to {BACKUP}")

    # Step 1: Update cell text using openpyxl (preserves VBA)
    update_cell_text()

    # Step 2: Inject drawing shapes AFTER openpyxl save
    # (openpyxl strips unknown elements, so drawing must go last)
    tmp = WORKBOOK + ".tmp"
    patch_workbook(WORKBOOK, tmp)
    os.replace(tmp, WORKBOOK)

    print(f"Patched {WORKBOOK} with 5 choice buttons (btnChoice1-5)")
    print("Each button calls ChoiceClicked_N macro on click.")
    print()
    print("Buttons are positioned at rows 25-29, columns B-C on the Game sheet.")


if __name__ == "__main__":
    main()
