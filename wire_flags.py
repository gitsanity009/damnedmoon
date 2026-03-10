#!/usr/bin/env python3
"""
wire_flags.py — Wire all 160 flags into scenes.

Part 1: Fill SetByScene for flags already referenced via FLAG_SET in tbl_Scenes.
Part 2: Add FLAG_SET effects to scenes for the 44 flags not yet wired.
Part 3: Add FLAG requirements to gate choices where narratively appropriate.
"""

import re
import openpyxl

WORKBOOK = "DamnedMoonv2.xlsm"

# ── Part 2: Manual wiring for the 44 flags not yet set by any scene ──────────
# Format: flag_name -> (scene_id, placement, column_hint)
#   placement: "on_enter", "on_exit", "c1_eff", "c2_eff", "c3_eff", "c4_eff", "c5_eff"

MANUAL_WIRING = {
    # -- Harlan relationship flags: set based on Harlan walk conversation outcomes --
    "HARLAN_REL_ASSET": ("SCN_003E_HARLAN_WALK", "c1_eff"),      # Walk into town (cooperative path)
    "HARLAN_REL_UNKNOWN": ("SCN_003_HARLAN", "c2_eff"),          # Keep it vague
    "HARLAN_REL_THREAT": ("SCN_003_HARLAN", "c3_eff"),           # Who's asking?
    "HARLAN_TOLD_ABOUT_MARIE": ("SCN_MARIE_HARLAN_Q", "c3_eff"), # Seemed like a harmless old man

    # -- Calhoun escalation flags --
    "CALHOUN_PROVOKED_1": ("SCN_DAY2_CALHOUN", "c2_eff"),       # Look him in the eye
    "CALHOUN_INCIDENT_1": ("SCN_CALHOUN_C2", "c1_eff"),         # Keep walking (files complaint anyway)
    "CALHOUN_METHOD_KNOWN": ("SCN_CALHOUN_C3", "c2_eff"),       # Acknowledge and explain
    "CALHOUN_PREPARING": ("SCN_CALHOUN_C4", "c2_eff"),          # Absorb the cost
    "CALHOUN_OWES_MARCUS": ("SCN_CALHOUN_C4", "c4_eff"),        # Confront directly — human to human
    "CHAINS_ACTIVE": ("SCN_CALHOUN_C4", "c3_eff"),              # Let the wolf surface (RAGE path)
    "CALHOUN_TAKEN": ("SCN_NIGHT30", "on_enter"),               # Pack moves — Calhoun was taken

    # -- Pack/Cord scenes --
    "PACK_CAMP_FOUND": ("SCN_NIGHT20_CURTIS", "on_enter"),       # Entering pack camp
    "CORD_SEEN": ("SCN_CURTIS_LISTEN", "on_enter"),              # Curtis explains = Cord visible
    "CORD_ACKNOWLEDGED_MARCUS": ("SCN_PACK_JOIN", "c1_eff"),     # Continue (joining the pack)
    "CORD_RESPECTS_RESISTANCE": ("SCN_CURTIS_LISTEN", "c2_eff"), # Refuse — walk away
    "CURTIS_WAS_CHOSEN": ("SCN_TAKODA_TALK", "on_enter"),        # Takoda reveals this

    # -- Darnell flags --
    "DARNELL_CALL_RECEIVED": ("SCN_DAY3_EVENING", "on_enter"),   # Darnell calls
    "DARNELL_IN_TOWN": ("SCN_DARNELL_INN", "on_enter"),          # Darnell arrives

    # -- Teen flags --
    "DELLA_TRUSTS_MARCUS": ("SCN_TEEN_RAPPORT", "c1_eff"),       # Come back tomorrow
    "HECTOR_SAW_SOMETHING": ("SCN_TEEN_HECTOR_WIN", "c2_eff"),   # Take the knife (wolf showed)
    "TEENS_HAVE_PACKAGE": ("SCN_TEEN_BARN", "on_enter"),         # Teens already have it
    "PACKAGE_LOST": ("SCN_DARNELL_INN", "c3_eff"),               # Honest — don't have it yet

    # -- Marie flags --
    "MARIE_GIFT_RECEIVED": ("SCN_MARIE_OFFER", "c1_eff"),        # Take her hand
    "MARIE_WITHDRAWN": ("SCN_PACK_JOIN", "on_enter"),             # Marie withdraws when you join pack
    "MARIE_ATTACKED": ("SCN_NIGHT30", "c3_eff"),                  # Marie's plan chosen (Marie path)
    "MARIE_EXPELLED": ("SCN_N30_CLEANSING", "on_enter"),          # Cleansing expels Marie's influence
    "COMBAT_SYSTEM_UNLOCKED": ("SCN_MARIE_OFFER", "c1_eff"),     # Marie housing = combat opens
    "GODDESS_THREAD_ACTIVE": ("SCN_HARLAN_RETURN", "c1_eff"),    # Share everything → Dominion activates

    # -- Dominion/Harlan --
    "DOMINION_CALL_MADE": ("SCN_HARLAN_RETURN", "c1_eff"),       # Share everything
    "CONTAINED_THREAD": ("SCN_HARLAN_RETURN", "c2_eff"),         # Share some — hold back

    # -- Investigation flags --
    "SMOKE_IN_NORTH_WOODS": ("SCN_NORTH_ROAD_V2", "on_enter"),   # Going deeper reveals smoke
    "JACOB_COMPASS_FOUND": ("SCN_NORTH_ROAD_V2", "c2_eff"),      # Head back (found compass on trail)
    "SWAMP_GEOGRAPHY_KNOWN": ("SCN_MARIE_APPROACH", "on_enter"),  # Walking swamp path = know layout
    "PRUITT_OFFICE_SEEN": ("SCN_004_PRUITT_CORDIAL", "c1_eff"),  # Take his hand (invited in)

    # -- Pruitt advanced flags --
    "PRUITT_FULL_DISCLOSURE": ("SCN_HARLAN_RETURN", "c1_eff"),   # Share everything (Pruitt path)
    "PRUITT_INVESTIGATING_HARLAN": ("SCN_HARLAN_RETURN", "c3_eff"),  # Ask who he really is
    "PRUITT_TAKEN": ("SCN_NIGHT30", "on_enter"),                  # Pack moves — Pruitt taken

    # -- Hattie --
    "HATTIE_DEAD": ("SCN_DARNELL_HOSTAGE", "c2_eff"),            # Let wolf surface (Darnell kills Hattie)
    "WOLFSBANE_OBTAINED": ("SCN_N30_CLEANSING", "c1_eff"),       # Place final marker (Hattie's gift)

    # -- Blackout progression --
    "BLACKOUT_EVENT_2": ("SCN_BLACKOUT", "c1_eff"),               # Wake up (second blackout tracked)
    "BLACKOUT_EVENT_3": ("SCN_BLACKOUT", "c1_eff"),               # Wake up (third blackout tracked)

    # -- Ending flags for missing endings --
    "ENDING_CHAINS": ("SCN_CALHOUN_C4", "c3_eff"),              # Wolf surfaces → chains ending path
    "ENDING_GODDESS_COMPANION": ("SCN_N30_MARIE_PLAN", "on_enter"),  # Marie plan = companion path
    "ENDING_GODDESS_PET": ("SCN_END_GENERALS_ARMY", "on_enter"),     # General's army = pet path
}

# ── Part 3: Additional FLAG requirements to gate choices ────────────────────
# Format: (scene_id, choice_num) -> requirement_to_add
REQUIREMENT_WIRING = {
    # Gate Marsh visit by having found the compass on north road
    # Already has: ("SCN_MARSH_VISIT", 1) -> FLAG:JACOB_COMPASS_FOUND  ✓

    # Gate Calhoun C3 accusation scene — needs teens to have package
    # Already has: ("SCN_CALHOUN_C3", 1) -> FLAG:TEENS_HAVE_PACKAGE  ✓

    # Gate Calhoun C4 breaking point — Darnell must be in town
    # Already has: ("SCN_CALHOUN_C4", 1) -> FLAG:DARNELL_IN_TOWN  ✓

    # Gate Harlan return — must have walked with Harlan
    ("SCN_HARLAN_RETURN", 1): "FLAG:HARLAN_WALKED_WITH",

    # Gate Cade E3 confession — must have seen something
    # Already has: ("SCN_CADE_E3", 1) -> FLAG:CADE_SAW_SOMETHING  ✓

    # Gate Curtis sighting — north road must be familiar
    ("SCN_CURTIS_SIGHTING", 1): "FLAG:NORTH_ROAD_FAMILIAR",

    # Gate Marie housing choices — need to have met Marie
    ("SCN_MARIE_OFFER", 1): "FLAG:MARIE_FORMALLY_MET",

    # Gate pack join based on Curtis alive
    ("SCN_PACK_JOIN", 1): "FLAG:CURTIS_ALIVE_ALLIED|FLAG:PACK_MEMBER",

    # Gate Darnell hostage — teens must have been contacted
    ("SCN_DARNELL_HOSTAGE", 2): "FLAG:TEENS_FIRST_CONTACT",
    ("SCN_DARNELL_HOSTAGE", 3): "FLAG:TEENS_FIRST_CONTACT",

    # Gate cleansing — must have met Takoda
    ("SCN_N30_CLEANSING", 1): "FLAG:TAKODA_MET",

    # Gate north road V3 — smoke must have been seen
    ("SCN_NORTH_ROAD_V3", 1): "FLAG:SHACK_FOUND",
}


def load_workbook():
    print(f"Loading {WORKBOOK}...")
    return openpyxl.load_workbook(WORKBOOK, keep_vba=True)


def get_scene_row_map(ws):
    """Build SceneID -> row mapping."""
    mapping = {}
    for r in range(2, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid:
            mapping[sid] = r
    return mapping


def part1_fill_set_by_scene(wb):
    """Scan tbl_Scenes for FLAG_SET references and fill tbl_Flags.SetByScene."""
    ws_scenes = wb["tbl_Scenes"]
    ws_flags = wb["tbl_Flags"]

    # Build flag name -> row mapping for tbl_Flags
    flag_rows = {}
    for r in range(2, ws_flags.max_row + 1):
        name = ws_flags.cell(r, 1).value
        if name:
            flag_rows[name] = r

    # Scan all scene cells for FLAG_SET references
    flag_to_scenes = {}
    for r in range(2, ws_scenes.max_row + 1):
        sid = ws_scenes.cell(r, 1).value
        if not sid:
            continue
        # Check OnEnter(27), OnExit(28), C1-C5 effects (10, 14, 18, 22, 26)
        for c in [27, 28, 10, 14, 18, 22, 26]:
            val = ws_scenes.cell(r, c).value
            if val and "FLAG_SET:" in str(val):
                for m in re.finditer(r"FLAG_SET:(\w+)", str(val)):
                    fname = m.group(1)
                    if fname not in flag_to_scenes:
                        flag_to_scenes[fname] = []
                    if sid not in flag_to_scenes[fname]:
                        flag_to_scenes[fname].append(sid)

    # Fill SetByScene (column 3) for flags that have it empty
    updated = 0
    for fname, scenes in sorted(flag_to_scenes.items()):
        if fname in flag_rows:
            row = flag_rows[fname]
            current = ws_flags.cell(row, 3).value
            if not current:
                # Use first scene that sets it (primary setter)
                ws_flags.cell(row, 3).value = scenes[0]
                updated += 1

    print(f"  Part 1: Filled SetByScene for {updated} flags")
    return flag_to_scenes


def append_effect(ws, row, col, new_effect):
    """Append a pipe-separated effect to an existing cell value."""
    current = ws.cell(row, col).value
    if current:
        current_str = str(current)
        # Don't add duplicate
        if new_effect in current_str:
            return False
        ws.cell(row, col).value = current_str + "|" + new_effect
    else:
        ws.cell(row, col).value = new_effect
    return True


def placement_to_col(placement):
    """Convert placement string to column number."""
    mapping = {
        "on_enter": 27,
        "on_exit": 28,
        "c1_eff": 10,
        "c2_eff": 14,
        "c3_eff": 18,
        "c4_eff": 22,
        "c5_eff": 26,
    }
    return mapping[placement]


def part2_wire_unset_flags(wb, existing_flag_scenes):
    """Add FLAG_SET effects to scenes for the 44 flags not yet wired."""
    ws_scenes = wb["tbl_Scenes"]
    ws_flags = wb["tbl_Flags"]
    scene_rows = get_scene_row_map(ws_scenes)

    # Build flag name -> row mapping
    flag_rows = {}
    for r in range(2, ws_flags.max_row + 1):
        name = ws_flags.cell(r, 1).value
        if name:
            flag_rows[name] = r

    added = 0
    skipped = []

    for flag_name, (scene_id, placement) in sorted(MANUAL_WIRING.items()):
        # Skip if flag is already set somewhere in scenes
        if flag_name in existing_flag_scenes:
            skipped.append(f"  SKIP (already set): {flag_name}")
            continue

        if scene_id not in scene_rows:
            skipped.append(f"  SKIP (scene not found): {flag_name} -> {scene_id}")
            continue

        row = scene_rows[scene_id]
        col = placement_to_col(placement)
        effect = f"FLAG_SET:{flag_name}"

        if append_effect(ws_scenes, row, col, effect):
            added += 1
            # Also fill SetByScene in tbl_Flags
            if flag_name in flag_rows:
                ws_flags.cell(flag_rows[flag_name], 3).value = scene_id
            print(f"    + {flag_name:45s} -> {scene_id} ({placement})")

    print(f"  Part 2: Added FLAG_SET for {added} flags to scenes")
    if skipped:
        for s in skipped:
            print(s)


def part3_add_requirements(wb):
    """Add FLAG requirements to gate choices where narratively appropriate."""
    ws_scenes = wb["tbl_Scenes"]
    scene_rows = get_scene_row_map(ws_scenes)

    added = 0
    for (scene_id, choice_num), req in sorted(REQUIREMENT_WIRING.items()):
        if scene_id not in scene_rows:
            print(f"  SKIP (scene not found): {scene_id}")
            continue

        row = scene_rows[scene_id]
        # Requirement columns: C1=9, C2=13, C3=17, C4=21, C5=25
        col = 9 + (choice_num - 1) * 4

        current = ws_scenes.cell(row, col).value
        if current:
            current_str = str(current)
            if req in current_str:
                continue
            ws_scenes.cell(row, col).value = current_str + "|" + req
        else:
            ws_scenes.cell(row, col).value = req

        added += 1
        print(f"    + {scene_id} C{choice_num} req: {req}")

    print(f"  Part 3: Added {added} requirements to choices")


def verify_integrity(wb):
    """Verify all flags have SetByScene and no orphan references."""
    ws_flags = wb["tbl_Flags"]
    ws_scenes = wb["tbl_Scenes"]

    # Check all flags have SetByScene
    missing = []
    total = 0
    for r in range(2, ws_flags.max_row + 1):
        name = ws_flags.cell(r, 1).value
        if not name:
            continue
        total += 1
        scene = ws_flags.cell(r, 3).value
        if not scene:
            missing.append(name)

    # Collect all scene IDs
    scene_ids = set()
    for r in range(2, ws_scenes.max_row + 1):
        sid = ws_scenes.cell(r, 1).value
        if sid:
            scene_ids.add(sid)

    # Check SetByScene references are valid
    invalid = []
    for r in range(2, ws_flags.max_row + 1):
        name = ws_flags.cell(r, 1).value
        scene = ws_flags.cell(r, 3).value
        if name and scene and scene not in scene_ids:
            invalid.append((name, scene))

    print(f"\n  Verification:")
    print(f"    Total flags: {total}")
    print(f"    Flags with SetByScene: {total - len(missing)}")
    print(f"    Flags missing SetByScene: {len(missing)}")
    if missing:
        for m in missing:
            print(f"      - {m}")
    print(f"    Invalid SetByScene refs: {len(invalid)}")
    if invalid:
        for name, scene in invalid:
            print(f"      - {name} -> {scene} (not found)")

    return len(missing), len(invalid)


def main():
    wb = load_workbook()

    print("\n── Part 1: Fill SetByScene from existing scene effects ──")
    existing = part1_fill_set_by_scene(wb)

    print("\n── Part 2: Wire unset flags into scenes ──")
    part2_wire_unset_flags(wb, existing)

    print("\n── Part 3: Add FLAG requirements to gate choices ──")
    part3_add_requirements(wb)

    print("\n── Verification ──")
    missing, invalid = verify_integrity(wb)

    print(f"\nSaving {WORKBOOK}...")
    wb.save(WORKBOOK)
    print("Done!")

    if missing > 0 or invalid > 0:
        print(f"\n⚠ {missing} flags still missing SetByScene, {invalid} invalid refs")
    else:
        print("\n✓ All flags wired successfully!")


if __name__ == "__main__":
    main()
