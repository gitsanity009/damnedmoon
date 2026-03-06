#!/usr/bin/env python3
"""
build_game_state.py — Creates the hidden GameState sheet in the workbook.

This sheet stores runtime state: stats, flags, scene history, and action log.
The VBA modState module reads/writes this sheet at runtime.

Run:  python3 build_game_state.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKBOOK = "BloodMoonProtocol_RPG.xlsx"

# ── Default stats matching modState.bas DEFAULT_STATS ──
DEFAULT_STATS = {
    "HP": 100, "MaxHP": 100,
    "Humanity": 100, "MaxHumanity": 100,
    "Rage": 0, "MaxRage": 100,
    "Hunger": 0, "MaxHunger": 100,
    "ScentLevel": 0, "Suspicion": 0,
    "SilverItems": 0, "MoonPhase": 5,
    "TimeOfDay": 347, "TransformStage": 0,
}

# ── Layout constants (must match modState.bas) ──
STAT_COL = 1    # A-B
FLAG_COL = 4    # D-E
HIST_COL = 7    # G-J
LOG_COL  = 12   # L-O
META_COL = 17   # Q-S

HDR_ROW  = 2
DATA_ROW = 3


def build_game_state(wb):
    """Create (or replace) the GameState sheet."""
    if "GameState" in wb.sheetnames:
        del wb["GameState"]

    ws = wb.create_sheet("GameState")

    # ── Styles ──
    section_font = Font(name="Consolas", bold=True, size=11, color="FFFFFF")
    section_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_font = Font(name="Consolas", bold=True, size=10, color="CCCCCC")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    data_font = Font(name="Consolas", size=10)
    thin_border = Border(
        left=Side(style="thin", color="444444"),
        right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"),
        bottom=Side(style="thin", color="444444"),
    )

    def write_cell(row, col, value, font=data_font, fill=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font
        c.border = thin_border
        if fill:
            c.fill = fill
        return c

    # ── Section labels (row 1) ──
    for col, label in [(STAT_COL, "STATS"), (FLAG_COL, "FLAGS"),
                       (HIST_COL, "HISTORY"), (LOG_COL, "ACTION LOG"),
                       (META_COL, "META")]:
        write_cell(1, col, label, section_font, section_fill)

    # ── Stats section ──
    write_cell(HDR_ROW, STAT_COL, "StatName", header_font, header_fill)
    write_cell(HDR_ROW, STAT_COL + 1, "Value", header_font, header_fill)
    for i, (name, val) in enumerate(DEFAULT_STATS.items()):
        write_cell(DATA_ROW + i, STAT_COL, name)
        write_cell(DATA_ROW + i, STAT_COL + 1, val)

    # ── Flags section ──
    write_cell(HDR_ROW, FLAG_COL, "FlagName", header_font, header_fill)
    write_cell(HDR_ROW, FLAG_COL + 1, "Value", header_font, header_fill)
    # (empty — flags are set at runtime)

    # ── History section ──
    for offset, label in enumerate(["#", "SceneID", "Choice", "Timestamp"]):
        write_cell(HDR_ROW, HIST_COL + offset, label, header_font, header_fill)

    # ── Action Log section ──
    for offset, label in enumerate(["#", "Action", "Detail", "Timestamp"]):
        write_cell(HDR_ROW, LOG_COL + offset, label, header_font, header_fill)

    # ── Meta section ──
    write_cell(HDR_ROW, META_COL, "CurrentScene", header_font, header_fill)
    write_cell(HDR_ROW, META_COL + 1, "SaveSlot", header_font, header_fill)
    write_cell(HDR_ROW, META_COL + 2, "SaveTimestamp", header_font, header_fill)
    write_cell(DATA_ROW, META_COL, "TITLE")
    write_cell(DATA_ROW, META_COL + 1, "Auto")

    # ── Column widths ──
    widths = {
        "A": 18, "B": 12,       # Stats
        "D": 22, "E": 12,       # Flags
        "G": 6, "H": 28, "I": 10, "J": 20,  # History
        "L": 6, "M": 16, "N": 50, "O": 20,  # Log
        "Q": 16, "R": 12, "S": 20,           # Meta
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Hide the sheet (VeryHidden equivalent — needs VBA to unhide)
    ws.sheet_state = "hidden"

    print(f"  ✓ GameState sheet created with {len(DEFAULT_STATS)} default stats")
    print(f"    Sections: Stats (A-B) | Flags (D-E) | History (G-J) | Log (L-O) | Meta (Q-S)")
    return ws


def main():
    print(f"Loading {WORKBOOK}...")
    wb = openpyxl.load_workbook(WORKBOOK)

    build_game_state(wb)

    wb.save(WORKBOOK)
    print(f"\nSaved to {WORKBOOK}")
    print("GameState sheet is hidden (use VBA or Python to inspect).")


if __name__ == "__main__":
    main()
