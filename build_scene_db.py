#!/usr/bin/env python3
"""
build_scene_db.py — Extracts scene data from existing narrative sheets
and writes a structured SceneDB sheet into the workbook.

Run:  python3 build_scene_db.py
"""

import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from copy import copy

WORKBOOK = "BloodMoonProtocol_RPG.xlsx"

# ── Column layout for SceneDB ──────────────────────────────────────────
HEADERS = [
    ("A", "SceneID",          "Unique scene identifier (matches sheet name)"),
    ("B", "SceneTitle",       "Display title from the scene header"),
    ("C", "StoryText",        "Full narrative text (rows 8-20)"),
    ("D", "HP",               "HP value displayed when entering scene"),
    ("E", "Humanity",         "Humanity value displayed when entering scene"),
    ("F", "MoonPhase",        "Moon phase indicator"),
    ("G", "ChoicePrompt",     "Prompt text above choices (row 22)"),
    ("H", "ChoiceA_Label",    "Choice A short label"),
    ("I", "ChoiceA_Desc",     "Choice A description"),
    ("J", "ChoiceA_Next",     "SceneID that Choice A links to"),
    ("K", "ChoiceB_Label",    "Choice B short label"),
    ("L", "ChoiceB_Desc",     "Choice B description"),
    ("M", "ChoiceB_Next",     "SceneID that Choice B links to"),
    ("N", "SceneType",        "choice | transition | ending | title"),
    ("O", "Warning",          "Warning text for dangerous choices"),
    ("P", "OnEnterEffects",   "Effects applied when entering scene (JSON)"),
    ("Q", "ConditionA",       "Condition required for Choice A (JSON)"),
    ("R", "ConditionB",       "Condition required for Choice B (JSON)"),
]


def parse_stat(text):
    """Extract numeric value from a stat display like 'HP  [████...] 85/100'."""
    if not text:
        return None
    m = re.search(r"(\d+)\s*/\s*\d+", str(text))
    return int(m.group(1)) if m else None


def parse_hyperlink_scene(cell):
    """Extract SceneID from a hyperlink target like '#SCENE_NAME!A1'."""
    if cell.hyperlink and cell.hyperlink.target:
        target = cell.hyperlink.target
        m = re.match(r"#?(\w+)!", target)
        if m:
            return m.group(1)
    return None


def classify_scene(scene_id, has_choice_b, story_text):
    """Determine scene type."""
    if scene_id == "TITLE":
        return "title"
    if scene_id.startswith("ENDING_"):
        return "ending"
    if has_choice_b:
        return "choice"
    return "transition"


def extract_scene(wb, sheet_name):
    """Extract structured data from one narrative sheet."""
    ws = wb[sheet_name]

    def cell(coord):
        return ws[coord].value

    # Story text: concatenate non-empty cells in rows 8-20, column A
    story_parts = []
    for r in range(8, 21):
        v = ws.cell(row=r, column=1).value
        if v:
            story_parts.append(str(v).strip())
    story_text = "\n\n".join(story_parts)

    # Stats
    hp = parse_stat(cell("A5"))
    humanity = parse_stat(cell("H5"))
    moon = cell("A6")

    # Choices
    choice_prompt = cell("A22")
    choice_a_label = cell("A23")
    choice_a_desc = cell("A24")
    choice_b_label = cell("H23")
    choice_b_desc = cell("H24")

    # Hyperlinks — check standard choice positions and fallbacks
    choice_a_next = parse_hyperlink_scene(ws["A27"])
    choice_b_next = parse_hyperlink_scene(ws["H27"])

    # Transition scenes: single button at various positions
    if not choice_a_next and not choice_b_next:
        for coord in ["C36", "D25", "D32"]:
            link = parse_hyperlink_scene(ws[coord])
            if link:
                choice_a_next = link
                choice_a_label = ws[coord].value
                break

    # Warning
    warning = cell("A29")

    # Scene type
    has_b = choice_b_label is not None
    scene_type = classify_scene(sheet_name, has_b, story_text)

    # Title special: pull story from different rows
    if sheet_name == "TITLE":
        story_parts = []
        for r in [8, 11]:
            v = ws.cell(row=r, column=1).value
            if v:
                story_parts.append(str(v).strip())
        story_text = "\n\n".join(story_parts)
        hp = 100
        humanity = 100

    return {
        "SceneID": sheet_name,
        "SceneTitle": cell("A4") or cell("A3") or sheet_name,
        "StoryText": story_text,
        "HP": hp,
        "Humanity": humanity,
        "MoonPhase": moon,
        "ChoicePrompt": choice_prompt,
        "ChoiceA_Label": choice_a_label,
        "ChoiceA_Desc": choice_a_desc,
        "ChoiceA_Next": choice_a_next,
        "ChoiceB_Label": choice_b_label,
        "ChoiceB_Desc": choice_b_desc,
        "ChoiceB_Next": choice_b_next,
        "SceneType": scene_type,
        "Warning": warning,
        "OnEnterEffects": "",
        "ConditionA": "",
        "ConditionB": "",
    }


def write_scene_db(wb, scenes):
    """Create or overwrite the SceneDB sheet."""
    if "SceneDB" in wb.sheetnames:
        del wb["SceneDB"]

    ws = wb.create_sheet("SceneDB", 0)  # Insert at front

    # ── Styles ──
    header_font = Font(name="Consolas", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    desc_font = Font(name="Consolas", italic=True, size=9, color="888888")
    data_font = Font(name="Consolas", size=10)
    thin_border = Border(
        left=Side(style="thin", color="444444"),
        right=Side(style="thin", color="444444"),
        top=Side(style="thin", color="444444"),
        bottom=Side(style="thin", color="444444"),
    )

    # ── Row 1: Headers ──
    for col_idx, (col_letter, name, desc) in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col_idx, value=name)
        c.font = header_font
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2: Descriptions ──
    for col_idx, (col_letter, name, desc) in enumerate(HEADERS, 1):
        c = ws.cell(row=2, column=col_idx, value=desc)
        c.font = desc_font
        c.border = thin_border
        c.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Data rows ──
    field_order = [h[1] for h in HEADERS]
    for row_idx, scene in enumerate(scenes, 3):
        for col_idx, field in enumerate(field_order, 1):
            val = scene.get(field, "")
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = data_font
            c.border = thin_border
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Column widths ──
    widths = {
        "A": 28, "B": 40, "C": 60, "D": 8, "E": 10,
        "F": 30, "G": 40, "H": 28, "I": 50, "J": 28,
        "K": 28, "L": 50, "M": 28, "N": 14, "O": 50,
        "P": 30, "Q": 30, "R": 30,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze top 2 rows
    ws.freeze_panes = "A3"

    # Auto-filter
    ws.auto_filter.ref = f"A1:R{2 + len(scenes)}"

    print(f"  ✓ SceneDB sheet created with {len(scenes)} scenes")


def main():
    print(f"Loading {WORKBOOK}...")
    wb = openpyxl.load_workbook(WORKBOOK)

    narrative_sheets = [s for s in wb.sheetnames if s != "SceneDB"]
    print(f"Found {len(narrative_sheets)} narrative sheets")

    scenes = []
    for name in narrative_sheets:
        scene = extract_scene(wb, name)
        scenes.append(scene)
        print(f"  • {scene['SceneID']:30s}  type={scene['SceneType']:12s}  → A={scene['ChoiceA_Next'] or '—':25s}  B={scene['ChoiceB_Next'] or '—'}")

    write_scene_db(wb, scenes)

    wb.save(WORKBOOK)
    print(f"\nSaved to {WORKBOOK}")
    print("SceneDB is the first sheet in the workbook.")


if __name__ == "__main__":
    main()
